import json
from pathlib import Path

from openpyxl import load_workbook

from integrations.demand_report.excel_writer import DEMAND_REPORT_COLUMNS, write_demand_report
from integrations.demand_report.extractor import extract_demand_item
from integrations.demand_report.keywords import KeywordPlan
from integrations.demand_report.models import DemandItem
from integrations.demand_report.runner import run_auto_demand_report
from integrations.demand_report.state import load_hashes, save_hashes


def test_extract_demand_item_prefers_comment_text():
    plan = KeywordPlan(keyword="装修 避坑", domain="装修", demand_word="避坑")
    item = extract_demand_item(
        {
            "comment_content": "求推荐靠谱的装修公司，预算不高怎么选？",
            "title": "装修记录",
            "note_url": "https://example.com/note/1",
            "nickname": "tester",
            "time": "2026-06-27",
        },
        plan,
        platform="xhs",
    )

    assert item is not None
    assert item.content_type == "评论"
    assert item.domain == "装修"
    assert item.demand_word == "避坑"
    assert item.content_hash


def test_state_round_trip(tmp_path: Path):
    path = tmp_path / "hashes.json"
    save_hashes({"a", "b"}, path)

    assert load_hashes(path) == {"a", "b"}


def test_write_demand_report_creates_xlsx(tmp_path: Path):
    item = DemandItem(
        title="求推荐靠谱的装修公司",
        raw_text="求推荐靠谱的装修公司，预算不高怎么选？",
        content_type="评论",
        platform="xhs",
        keyword="装修 避坑",
        domain="装修",
        demand_word="避坑",
        source_url="https://example.com/note/1",
        author="tester",
        publish_time="2026-06-27",
        collected_at="2026-06-27 12:00:00",
        content_hash="hash1",
    )

    output = write_demand_report([item], output_dir=tmp_path, now_text="2026-06-27-1200")

    assert output.exists()
    wb = load_workbook(output)
    ws = wb.active
    assert [cell.value for cell in ws[1]] == DEMAND_REPORT_COLUMNS
    assert ws["A2"].value == "求推荐靠谱的装修公司"


def test_run_auto_demand_report_dry_run_does_not_execute(tmp_path: Path):
    result = run_auto_demand_report(
        platforms=["xhs"],
        keyword_count=2,
        dry_run=True,
        output_dir=tmp_path,
        executor=lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("should not execute")),
        send_summary=lambda **_kwargs: True,
    )

    assert result.dry_run is True
    assert result.stats.total_keywords == 2
    assert result.excel_path is None


def test_run_auto_demand_report_reads_exports_and_writes_excel(tmp_path: Path):
    export_dir = tmp_path / "data" / "xhs" / "jsonl"
    export_dir.mkdir(parents=True)
    export_file = export_dir / "items.jsonl"
    export_file.write_text(
        json.dumps(
            {
                "comment_content": "求推荐好用的AI工具，预算有限怎么选？",
                "note_url": "https://example.com/note/2",
                "nickname": "tester",
            },
            ensure_ascii=False,
        )
        + "\n",
        encoding="utf-8",
    )

    sent = {}
    result = run_auto_demand_report(
        platforms=["xhs"],
        keyword_count=1,
        dry_run=False,
        data_dir=tmp_path / "data",
        output_dir=tmp_path / "reports",
        state_path=tmp_path / "state.json",
        executor=lambda *_args, **_kwargs: None,
        send_summary=lambda **kwargs: sent.update(kwargs) or True,
    )

    assert result.excel_path is not None
    assert result.excel_path.exists()
    assert result.stats.new_items == 1
    assert "excel_path" in sent
