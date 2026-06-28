# -*- coding: utf-8 -*-
"""Excel writer for automated demand reports."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import DemandItem


DEMAND_REPORT_COLUMNS: List[str] = [
    "需求标题",
    "需求原文",
    "内容类型",
    "来源平台",
    "自动关键词",
    "领域",
    "需求词",
    "来源链接",
    "作者/昵称",
    "发布时间",
    "采集时间",
    "内容哈希",
    "备注",
]


def write_demand_report(
    items: Iterable[DemandItem],
    output_dir: Path = Path("output") / "demand_reports",
    now_text: str | None = None,
) -> Path:
    """Write demand items to an `.xlsx` report and return the output path."""

    item_list = list(items)
    if not item_list:
        raise ValueError("items must not be empty")

    output_dir.mkdir(parents=True, exist_ok=True)
    now_text = now_text or datetime.now().strftime("%Y-%m-%d-%H%M")
    output_path = output_dir / f"{now_text}-demand-report.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "需求线索"
    ws.append(DEMAND_REPORT_COLUMNS)

    for item in item_list:
        ws.append(
            [
                item.title,
                item.raw_text,
                item.content_type,
                item.platform,
                item.keyword,
                item.domain,
                item.demand_word,
                item.source_url,
                item.author,
                item.publish_time,
                item.collected_at,
                item.content_hash,
                item.remark,
            ]
        )

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = [22, 56, 12, 12, 18, 12, 12, 36, 16, 18, 20, 24, 18]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width

    wb.save(output_path)
    return output_path

